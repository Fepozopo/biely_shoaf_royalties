import openpyxl



def update_sales(artists, royalty_file, report_file):
    # Loop through artists
    for artist in artists:
        wb_royalty = openpyxl.load_workbook(royalty_file)
        ws_royalty = wb_royalty[f'{artist}']

        wb_report = openpyxl.load_workbook(report_file)
        ws_report = wb_report['Sheet1']

        # Define columns
        royalty_sku_col = 1 # 'A' column in the royalty file (1-based index)
        royalty_sales_col = 3 # 'C' column in the royalty file (1-based index)
        report_sku_col = 3 # 'C' column in the report (1-based index)
        report_sales_col = 6 # 'F' column in the report (1-based index)
        report_pointer = 1 # Pointer to the next row in the report

        print(f'Updating {artist}...')
        for row in range(3, ws_royalty.max_row + 1): # Loop through royalty rows
            royalty_sku = ws_royalty.cell(row=row, column=royalty_sku_col).value
            
            if royalty_sku is None:
                continue # Skip empty rows

            # Ensure royalty_sku is a string and strip leading/trailing whitespace
            royalty_sku = str(royalty_sku).strip()

            sku_found = False  # Initialize flag to track if SKU is found

            for row_report in range(report_pointer, ws_report.max_row + 1): # Loop through report rows
                report_sku = ws_report.cell(row=row_report, column=report_sku_col).value

                if report_sku is None:
                    continue # Skip empty rows

                # Ensure report_sku is a string and strip leading/trailing whitespace
                report_sku = str(report_sku).strip()

                if royalty_sku == report_sku:
                    report_sales_value = ws_report.cell(row=row_report, column=report_sales_col).value
                    
                    ws_royalty.cell(row=row, column=royalty_sales_col).value = report_sales_value # Update the royalty sales column with the report sales column
                    print(f"Updated SKU {royalty_sku} with sales value {report_sales_value}")
                    
                    report_pointer = row_report + 1 # Move the report pointer to the next row
                    sku_found = True # Mark the SKU as found
                    break # Move to the next royalty row

            # If SKU was not found in the report, set sales to 0
            if not sku_found:
                ws_royalty.cell(row=row, column=royalty_sales_col).value = 0
                print(f"SKU {royalty_sku} not found, setting sales to 0")

        # Save and close the workbook
        wb_royalty.save(royalty_file)
        wb_report.close()
        wb_royalty.close()


def update_royalty(artists, royalty_file):
    # Loop through artists
    for artist in artists:
        wb = openpyxl.load_workbook(royalty_file)
        ws = wb[f'{artist}']

        # Define columns
        sku_col = 1 # 'A' column in the royalty file (1-based index)
        balance_col = 2 # 'B' column in the royalty file (1-based index)
        sales_col = 3 # 'C' column in the royalty file (1-based index)
        rate_col = 4 # 'D' column in the royalty file (1-based index)
        earned_col = 5 # 'E' column in the royalty file (1-based index)
        new_balance_col = 7 # 'G' column in the royalty file (1-based index)

        print(f'Updating {artist}...')
        for row in range(3, ws.max_row + 1): # Loop through royalty rows
            sku = ws.cell(row=row, column=sku_col).value

            if sku is None:
                continue # Skip empty rows

            # Move the value in new_balance_col to balance_col
            ws.cell(row=row, column=balance_col).value = float(ws.cell(row=row, column=new_balance_col).value)
            print(f"Updated SKU {sku} with balance value {ws.cell(row=row, column=new_balance_col).value}")

            # Calculate the earned_col by multiplying sales_col by rate_col by (rate_col/100)
            ws.cell(row=row, column=earned_col).value = float(ws.cell(row=row, column=sales_col).value * (ws.cell(row=row, column=rate_col).value / 100))
            print(f"Updated SKU {sku} with earned value {ws.cell(row=row, column=earned_col).value}")

            # Calculate the new balance by subtracting earned_col from balance_col using the formula '=MAX(0, balance_col - earned_col)'
            ws.cell(row=row, column=new_balance_col).value = float(max(0, ws.cell(row=row, column=balance_col).value - ws.cell(row=row, column=earned_col).value))
            print(f"Updated SKU {sku} with new balance value {ws.cell(row=row, column=new_balance_col).value}")


        # Save and close the workbook
        wb.save(royalty_file)
        wb.close()
